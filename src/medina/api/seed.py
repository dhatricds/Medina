"""Seed the dashboard with training xlsx files on first startup."""
from __future__ import annotations

import json
import logging
import re
from datetime import datetime, timezone
from pathlib import Path

from openpyxl import load_workbook

logger = logging.getLogger(__name__)

TRAIN_DIR = Path(__file__).resolve().parents[3] / "train"
DASHBOARD_DIR = Path(__file__).resolve().parents[3] / "output" / "dashboard"


def _sanitize_id(name: str) -> str:
    """Turn a filename into a safe dashboard ID."""
    stem = Path(name).stem.replace("_inventory", "")
    return re.sub(r"[^a-zA-Z0-9_-]", "_", stem)[:60]


def _parse_xlsx(xlsx_path: Path) -> dict:
    """Parse a training xlsx into a ProjectData-compatible dict."""
    wb = load_workbook(xlsx_path, read_only=True, data_only=True)

    # --- Fixture Inventory sheet ---
    fixtures: list[dict] = []
    plan_codes: list[str] = []

    if "Fixture Inventory" in wb.sheetnames:
        ws = wb["Fixture Inventory"]
        rows = list(ws.iter_rows(values_only=True))
        if rows:
            headers = [str(h).strip() if h else "" for h in rows[0]]

            # Fixed spec columns end before the plan count columns
            spec_labels = {
                "Type", "Description", "Fixture Style", "Voltage",
                "Mounting", "Lumens", "CCT", "Dimming", "Max VA",
            }
            spec_end = 0
            for i, h in enumerate(headers):
                if h in spec_labels:
                    spec_end = i + 1

            # Plan columns: between spec_end and "Total"
            total_idx = None
            for i, h in enumerate(headers):
                if h.lower() == "total":
                    total_idx = i
                    break

            if total_idx and total_idx > spec_end:
                plan_codes = [headers[i] for i in range(spec_end, total_idx)]

            # Map header labels to model field names
            field_map = {
                "Type": "code",
                "Description": "description",
                "Fixture Style": "fixture_style",
                "Voltage": "voltage",
                "Mounting": "mounting",
                "Lumens": "lumens",
                "CCT": "cct",
                "Dimming": "dimming",
                "Max VA": "max_va",
            }

            for row in rows[1:]:
                if not row or not row[0]:
                    break  # Stop at empty rows (before Key Notes Summary)
                # Skip non-fixture rows
                code_val = str(row[0]).strip() if row[0] else ""
                if not code_val or code_val.lower() in ("key notes summary", "#"):
                    break

                fixture: dict = {
                    "code": "",
                    "description": "",
                    "fixture_style": "",
                    "voltage": "",
                    "mounting": "",
                    "lumens": "",
                    "cct": "",
                    "dimming": "",
                    "max_va": "",
                    "counts_per_plan": {},
                    "total": 0,
                }

                for col_idx, header in enumerate(headers):
                    if col_idx >= len(row):
                        break
                    val = row[col_idx]
                    if header in field_map:
                        fixture[field_map[header]] = str(val).strip() if val else ""
                    elif header in plan_codes:
                        fixture["counts_per_plan"][header] = int(val) if val else 0
                    elif header.lower() == "total":
                        fixture["total"] = int(val) if val else 0

                if fixture["code"]:
                    fixtures.append(fixture)

    # --- Key Notes Inventory sheet ---
    keynotes: list[dict] = []

    if "Key Notes Inventory" in wb.sheetnames:
        ws_kn = wb["Key Notes Inventory"]
        kn_rows = list(ws_kn.iter_rows(values_only=True))
        if kn_rows:
            kn_headers = [str(h).strip() if h else "" for h in kn_rows[0]]

            for row in kn_rows[1:]:
                if not row or not row[0]:
                    break
                kn_num = str(row[0]).strip() if row[0] else ""
                if not kn_num:
                    break

                kn: dict = {
                    "keynote_number": kn_num,
                    "keynote_text": str(row[1]).strip() if len(row) > 1 and row[1] else "",
                    "counts_per_plan": {},
                    "total": 0,
                    "fixture_references": [],
                }

                for col_idx, header in enumerate(kn_headers):
                    if col_idx >= len(row):
                        break
                    val = row[col_idx]
                    if header.lower() == "total":
                        kn["total"] = int(val) if val else 0
                    elif col_idx >= 2 and header.lower() != "total":
                        # Plan count column
                        if header and header not in ("Keynote #", "Keynote Text"):
                            kn["counts_per_plan"][header] = int(val) if val else 0

                keynotes.append(kn)

    # --- QA Report sheet ---
    qa_report = None
    if "QA Report" in wb.sheetnames:
        ws_qa = wb["QA Report"]
        qa_rows = list(ws_qa.iter_rows(values_only=True))
        overall_conf = 0.0
        stage_scores: dict[str, float] = {}
        warnings: list[str] = []

        stage_key_map = {
            "Sheet Index Discovery": "sheet_index",
            "Schedule Extraction": "schedule_extraction",
            "Fixture Counting": "fixture_counting",
            "Keynote Extraction": "keynote_extraction",
        }

        in_warnings = False
        for row in qa_rows:
            if not row:
                continue
            label = str(row[0]).strip() if row[0] else ""
            val = str(row[1]).strip() if len(row) > 1 and row[1] else ""

            if label == "Overall Confidence" and val:
                overall_conf = float(val.replace("%", "")) / 100.0
            elif label in stage_key_map and val:
                stage_scores[stage_key_map[label]] = float(val.replace("%", "")) / 100.0
            elif label == "Warnings":
                in_warnings = True
                continue
            elif label == "Recommendations":
                in_warnings = False
            elif in_warnings and label:
                warnings.append(label)

        qa_report = {
            "overall_confidence": overall_conf,
            "passed": overall_conf >= 0.95,
            "threshold": 0.95,
            "stage_scores": stage_scores,
            "warnings": warnings,
            "recommendations": [],
        }

    wb.close()

    # Build project name from filename
    stem = xlsx_path.stem.replace("_inventory", "")
    # Clean up DENIS names
    project_name = re.sub(r"\([\d]+\)", "", stem).strip()

    total_fixtures = sum(f["total"] for f in fixtures)

    project_data = {
        "project_name": project_name,
        "sheet_index": [],
        "lighting_plans": plan_codes,
        "schedule_pages": [],
        "fixtures": fixtures,
        "keynotes": keynotes,
        "summary": {
            "total_fixture_types": len(fixtures),
            "total_fixtures": total_fixtures,
            "total_lighting_plans": len(plan_codes),
            "total_keynotes": len(keynotes),
        },
        "qa_report": qa_report,
    }

    return project_data


def seed_dashboard() -> None:
    """Seed the dashboard directory with training xlsx files if not already done."""
    DASHBOARD_DIR.mkdir(parents=True, exist_ok=True)
    index_path = DASHBOARD_DIR / "index.json"

    if index_path.exists():
        logger.info("Dashboard already seeded (%s exists)", index_path)
        return

    if not TRAIN_DIR.exists():
        logger.warning("Training directory not found: %s", TRAIN_DIR)
        return

    xlsx_files = sorted(TRAIN_DIR.glob("*.xlsx"))
    if not xlsx_files:
        logger.warning("No xlsx files found in %s", TRAIN_DIR)
        return

    index: list[dict] = []
    now = datetime.now(timezone.utc).isoformat()

    for xlsx_path in xlsx_files:
        try:
            project_id = _sanitize_id(xlsx_path.name)
            project_data = _parse_xlsx(xlsx_path)

            # Save project JSON
            project_json_path = DASHBOARD_DIR / f"{project_id}.json"
            with open(project_json_path, "w") as f:
                json.dump(project_data, f, indent=2)

            # Copy xlsx for export
            import shutil
            dest_xlsx = DASHBOARD_DIR / f"{project_id}.xlsx"
            shutil.copy2(xlsx_path, dest_xlsx)

            # Build index entry
            qa = project_data.get("qa_report")
            entry = {
                "id": project_id,
                "name": project_data["project_name"],
                "approved_at": now,
                "fixture_types": project_data["summary"]["total_fixture_types"],
                "total_fixtures": project_data["summary"]["total_fixtures"],
                "keynote_count": project_data["summary"]["total_keynotes"],
                "plan_count": project_data["summary"]["total_lighting_plans"],
                "qa_score": qa["overall_confidence"] if qa else None,
                "qa_passed": qa["passed"] if qa else None,
            }
            index.append(entry)
            logger.info("Seeded dashboard project: %s", project_data["project_name"])

        except Exception:
            logger.exception("Failed to seed %s", xlsx_path.name)

    # Write index
    with open(index_path, "w") as f:
        json.dump(index, f, indent=2)

    logger.info("Dashboard seeded with %d projects", len(index))
