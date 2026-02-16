"""Excel workbook generation with dynamic per-plan columns."""

from __future__ import annotations

import logging
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

from medina.models import ExtractionResult, QAReport

logger = logging.getLogger(__name__)

SPEC_COLUMNS = [
    ("code", "Type"),
    ("description", "Description"),
    ("fixture_style", "Fixture Style"),
    ("voltage", "Voltage"),
    ("mounting", "Mounting"),
    ("lumens", "Lumens"),
    ("cct", "CCT"),
    ("dimming", "Dimming"),
    ("max_va", "Max VA"),
]

GREEN_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
YELLOW_FILL = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
RED_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
HEADER_FONT = Font(bold=True, size=11)
TITLE_FONT = Font(bold=True, size=14)


def _auto_width(ws, min_width: int = 10, max_width: int = 40) -> None:
    """Auto-size column widths based on content."""
    for col_cells in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col_cells[0].column)
        for cell in col_cells:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        width = min(max(max_len + 2, min_width), max_width)
        ws.column_dimensions[col_letter].width = width


def generate_fixture_sheet(
    wb: Workbook,
    result: ExtractionResult,
) -> None:
    """Create the Fixture Inventory sheet with dynamic plan columns."""
    ws = wb.active
    ws.title = "Fixture Inventory"

    plan_codes = result.plan_pages or []
    headers = [label for _, label in SPEC_COLUMNS]
    headers += plan_codes
    headers.append("Total")

    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = HEADER_FONT

    for row_idx, fixture in enumerate(result.fixtures, 2):
        for col_idx, (field, _) in enumerate(SPEC_COLUMNS, 1):
            ws.cell(
                row=row_idx,
                column=col_idx,
                value=getattr(fixture, field, ""),
            )

        for plan_idx, plan_code in enumerate(plan_codes):
            col = len(SPEC_COLUMNS) + plan_idx + 1
            count = fixture.counts_per_plan.get(plan_code, 0)
            ws.cell(row=row_idx, column=col, value=count)

        total_col = len(SPEC_COLUMNS) + len(plan_codes) + 1
        ws.cell(row=row_idx, column=total_col, value=fixture.total)

    if result.fixtures:
        last_row = len(result.fixtures) + 1
        last_col = get_column_letter(len(headers))
        table_ref = f"A1:{last_col}{last_row}"
        table = Table(displayName="LightingInventory", ref=table_ref)
        style = TableStyleInfo(
            name="TableStyleMedium9",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False,
        )
        table.tableStyleInfo = style
        ws.add_table(table)

    if result.keynotes:
        summary_start = len(result.fixtures) + 4
        ws.cell(
            row=summary_start,
            column=1,
            value="Key Notes Summary",
        ).font = TITLE_FONT

        row = summary_start + 1
        kn_headers = ["#", "Key Note Text", "Total"]
        for col_idx, h in enumerate(kn_headers, 1):
            ws.cell(row=row, column=col_idx, value=h).font = HEADER_FONT

        for kn in result.keynotes:
            row += 1
            ws.cell(row=row, column=1, value=str(kn.number))
            ws.cell(row=row, column=2, value=kn.text)
            ws.cell(row=row, column=3, value=kn.total)

    _auto_width(ws)


def generate_keynotes_sheet(
    wb: Workbook,
    result: ExtractionResult,
) -> None:
    """Create the Key Notes Inventory sheet."""
    ws = wb.create_sheet("Key Notes Inventory")
    plan_codes = result.plan_pages or []

    headers = ["Keynote #", "Keynote Text"]
    headers += plan_codes
    headers.append("Total")

    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = HEADER_FONT

    for row_idx, kn in enumerate(result.keynotes, 2):
        ws.cell(row=row_idx, column=1, value=str(kn.number))
        ws.cell(row=row_idx, column=2, value=kn.text)

        for plan_idx, plan_code in enumerate(plan_codes):
            col = 3 + plan_idx
            count = kn.counts_per_plan.get(plan_code, 0)
            ws.cell(row=row_idx, column=col, value=count)

        total_col = 3 + len(plan_codes)
        ws.cell(row=row_idx, column=total_col, value=kn.total)

    if result.keynotes:
        last_row = len(result.keynotes) + 1
        last_col = get_column_letter(len(headers))
        table_ref = f"A1:{last_col}{last_row}"
        table = Table(displayName="KeyNotesInventory", ref=table_ref)
        style = TableStyleInfo(
            name="TableStyleMedium9",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False,
        )
        table.tableStyleInfo = style
        ws.add_table(table)

    _auto_width(ws)


def generate_qa_sheet(wb: Workbook, report: QAReport) -> None:
    """Create the QA Report sheet with color-coded confidence scores."""
    ws = wb.create_sheet("QA Report")

    ws.cell(row=1, column=1, value="MEDINA QA REPORT").font = TITLE_FONT

    status = "PASSED" if report.passed else "FAILED"
    ws.cell(row=3, column=1, value="Overall Confidence")
    conf_cell = ws.cell(
        row=3, column=2, value=f"{report.overall_confidence:.1%}"
    )
    ws.cell(row=3, column=3, value=status)

    if report.overall_confidence >= 0.95:
        conf_cell.fill = GREEN_FILL
    elif report.overall_confidence >= 0.80:
        conf_cell.fill = YELLOW_FILL
    else:
        conf_cell.fill = RED_FILL

    ws.cell(row=4, column=1, value="Threshold")
    ws.cell(row=4, column=2, value=f"{report.threshold:.0%}")

    row = 6
    ws.cell(row=row, column=1, value="Stage Scores").font = HEADER_FONT
    row += 1
    stage_labels = {
        "sheet_index": "Sheet Index Discovery",
        "schedule_extraction": "Schedule Extraction",
        "fixture_counting": "Fixture Counting",
        "keynote_extraction": "Keynote Extraction",
    }
    for key, label in stage_labels.items():
        score = report.stage_scores.get(key, 0.0)
        ws.cell(row=row, column=1, value=label)
        score_cell = ws.cell(row=row, column=2, value=f"{score:.1%}")
        if score >= 0.95:
            score_cell.fill = GREEN_FILL
        elif score >= 0.80:
            score_cell.fill = YELLOW_FILL
        else:
            score_cell.fill = RED_FILL
        row += 1

    if report.warnings:
        row += 1
        ws.cell(row=row, column=1, value="Warnings").font = HEADER_FONT
        for w in report.warnings:
            row += 1
            ws.cell(row=row, column=1, value=w)

    if report.recommendations:
        row += 1
        ws.cell(
            row=row, column=1, value="Recommendations"
        ).font = HEADER_FONT
        for r in report.recommendations:
            row += 1
            ws.cell(row=row, column=1, value=r)

    if report.fixture_results:
        row += 2
        ws.cell(
            row=row, column=1, value="Fixture QA Details"
        ).font = HEADER_FONT
        row += 1
        for h_idx, h in enumerate(
            ["Code", "Confidence", "Flags", "Notes"], 1
        ):
            ws.cell(row=row, column=h_idx, value=h).font = HEADER_FONT

        for item in report.fixture_results:
            row += 1
            ws.cell(row=row, column=1, value=item.item_code)
            c = ws.cell(row=row, column=2, value=f"{item.confidence:.0%}")
            if item.confidence >= 0.95:
                c.fill = GREEN_FILL
            elif item.confidence >= 0.80:
                c.fill = YELLOW_FILL
            else:
                c.fill = RED_FILL
            flags = ", ".join(f.value for f in item.flags)
            ws.cell(row=row, column=3, value=flags)
            ws.cell(row=row, column=4, value=item.notes)

    _auto_width(ws)


def write_excel(
    result: ExtractionResult,
    output_path: str | Path,
) -> Path:
    """Generate the complete Excel workbook."""
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    generate_fixture_sheet(wb, result)
    generate_keynotes_sheet(wb, result)

    if result.qa_report:
        generate_qa_sheet(wb, result.qa_report)

    wb.save(str(output_path))
    logger.info("Excel workbook saved to %s", output_path)
    return output_path
